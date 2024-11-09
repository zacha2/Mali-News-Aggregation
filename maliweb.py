import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import time

link_international = "https://www.maliweb.net/category/international/page/"



def scrape_pages(link):
    #loop for number of pages to scrape
    articles = []
    for page in range(1, 580):
        
        url = f'{link}{page}'

        """used headers to prevent the website from flagging the script

        I had to use a VPN for maliweb because the website had already blacklisted my IP address"""

        headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'}

        result = requests.get(url, headers=headers)
        scraper = BeautifulSoup(result.text, 'html.parser')

        print(page)

        #article_containers has every article per page of articles
        page_container = scraper.find('div', class_="td-main-content")
        article_containers = page_container.find_all('div', class_="td_module_wrap")

        #gets article info from every article container in page
        for article_container in article_containers:
        
            article_url = article_container.find('h3', class_="entry-title").find('a', href=True)['href']
            article_name = article_container.find('h3', class_='entry-title').text
            article_date = article_container.find('time', class_="entry-date").text
            
            #Splits the article date into monthly segments
            article_month = article_date.split(' ')[1] + article_date.split(' ')[2]

        
            article = {'name': article_name,
                       'month': article_month, 
                       'keywords': {'Russie': 0, 'Russe': 0, 'Poutine': 0, 'Sputnik': 0, 'RT': 0, 'New Eastern Outlook': 0, 'Tass': 0,},
                       }
            
            #Scraper for each individual article
            article_result = requests.get(article_url)
            article_soup = BeautifulSoup(article_result.text, 'html.parser')

            article_contents = article_soup.find('div', class_='td-post-content').text

            #looks for keyword in article and adds it to that articles dictionary
            for keyword in article['keywords']:
                if keyword in article_contents:
                    article['keywords'][keyword] += 1
            
            articles.append(article)
            
            #Makes the script wait 1 second as an aditional measure to make sure it doesn't get flagged
            time.sleep(1)

    #keeps track of all the months found in articles
    monthly_keywords = {}

    for article in articles:
        month = article['month']
        keywords = article['keywords']
        
        if month not in monthly_keywords:
            monthly_keywords[month] = {}
        
        """Loops through dictionary of monthly_keywords and adds keyword counts to monthly_keywords

            if the keyword isn't already in the montly_keywords, it is added

            the value of the keyword is then added to monthly_keywords[month][keyword]
        """
        for keyword, count in keywords.items():
            if keyword not in monthly_keywords[month]:
                monthly_keywords[month][keyword] = 0
            monthly_keywords[month][keyword] += count

    
    #Creates an openpyxl workbook for articles data
    wb = Workbook()
    ws = wb.active

    
    ws['B1'] = 'Russie'
    ws['C1'] = 'Russe'
    ws['D1'] = 'Poutine'
    ws['E1'] = 'Sputnik'
    ws['F1'] = 'RT'
    ws['G1'] = 'New Eastern Outlook'
    ws['H1'] = 'Tass'

    #Loop that organizes monthly keyword list and places them on a seperate row per month.
    for month, results  in monthly_keywords.items():
        item_list = []
        print(month)
        item_list.append(month)
        for key, value in results.items():
            item_list.append(value)
        ws.append(item_list)
    wb.save('maliweb_data_pg1-200.xlsx')




scrape_pages(link_international)

