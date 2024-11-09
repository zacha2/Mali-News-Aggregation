import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import time

link_international = "https://malijet.com/actualite_internationale/?page="

def scrape_pages(link):
    articles = []

    #for loop for each page to be scraped
    for page in range(1, 276):
        url = f'{link}{page}'

        headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'}

        result = requests.get(url, headers=headers)
        scraper = BeautifulSoup(result.text, 'html.parser')

        #page_container has every article per page of articles
        page_container = scraper.find('div', id="v_container")
        article_list = page_container.find_all('div', class_="col-md-6")

        print(page)

        #This for loop collects basic information for each article
        for article_info in article_list:
            article_date_item = article_info.find('div', class_="bas_articles").find('h5').find_next_sibling('h5').text
            article_date = article_date_item.split(' ')[2] + article_date_item.split(' ')[3]

            article_url = article_info.find('h5', class_="card-title").find('a', href=True)['href']
            
            article_dictionary = {"url": article_url, "date": article_date, 
                                  "keywords": {'Russie': 0, 'Russe': 0, 'Poutine': 0, 'Sputnik': 0, 'RT': 0, 'New Eastern Outlook': 0, 'Tass': 0,},
                       }
            
            #Scraper for each individual article
            article_request = requests.get(article_url)
            article_scraper = BeautifulSoup(article_request.text, "html.parser")


            text_container = article_scraper.find('div', id="card_img_jt")

            #Added if statement since the script would occaisonally say there was no text_container
            if text_container.text != None:
                for keyword in article_dictionary["keywords"]:
                    if keyword in text_container.text:
                        article_dictionary["keywords"][keyword] += 1

                articles.append(article_dictionary)

                time.sleep(1)

            

        months = {}
        
        #goes through every article and turns every month found into a dictionary in the months dictionary
        for article in articles:
            date = article["date"]
            keywords = article["keywords"]

            if date not in months:
                months[date] = {}

            """Loops through dictionary of monthly_keywords and adds keyword counts to monthly_keywords

            if the keyword isn't already in the montly_keywords, it is added

            the value of the keyword, count is then added to monthly_keywords[month][keyword]
        """
            for keyword, count in keywords.items():

                if keyword not in months[date]:
                    months[date][keyword] = 0
                
                #updates keyword count for each month
                months[date][keyword] += count
        
        print(months)

        wb = Workbook()
        ws = wb.active

    ws['B1'] = 'Russie'
    ws['C1'] = 'Russe'
    ws['D1'] = 'Poutine'
    ws['E1'] = 'Sputnik'
    ws['F1'] = 'RT'
    ws['G1'] = 'New Eastern Outlook'
    ws['H1'] = 'Tass'

    #organizes monthly keywords and places them on a seperate row per month
    for month, results  in months.items():
        item_list = []
        
        item_list.append(month)
        for key, value in results.items():
            item_list.append(value)
        ws.append(item_list)
    wb.save('malijet_data_pg1-276.xlsx')



scrape_pages(link_international)

